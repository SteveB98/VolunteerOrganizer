import openpyxl as pyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl import Workbook
import ScheduleUI
import SheetFormat
from ScheduleUI import parse_args
from SheetFormat import title_font, bolded_font, supervisor_font, align
"""
Program: ScheduleOrganizer.py
Purpose: Handles backend openpyxl processing of Excel Workbooks
and spreadsheets. Creates or savesr new workbook with user input from ScheduleUI.py
with visual grids and conenciding shift sheets
"""
#FUNCTIONS:
#Parameters: Cells ([date,venue,position,time]), Sheet (Sheet Obj), Ref_Sheet (Reference Sheet Title), Vol_Cell (Value of Volunteer Cell)
#Purpose: With a passed through cell location, venue and shift list sheets, this function compiles direct reference Excel formulas from cells from a given
#venue grid to the shift list. Taking the date, venue, work position, and time details for each shift & matching it with it's volunteer name. 
def insert_sheet_list(cells,sheet,ref_sheet,vol_cell):
        sheet_call = ref_sheet.title
        space_formula = '&", "&'
        #Apostrophe in Sheet title check, replace with double apostrophe if so
        if sheet_call.find("'") != -1: sheet_call = sheet_call.replace("'","\'\'")
        #Space check in sheet title, insert single quotes around title if so
        if sheet_call.find(" ") != -1: sheet_call = "'"+sheet_call+"'"
        sheet_call = sheet_call+"!"
        #Finding inital cell for position cells, does the same to time cells if universal shift time is required
        position_merge = [r for r in ref_sheet.merged_cells.ranges if cells[2] in r][0].start_cell.coordinate
        if args.UniversalShiftTime == True: time_value = [r for r in ref_sheet.merged_cells.ranges if cells[3] in r][0].start_cell.coordinate
        else: time_value = cells[3]
        #Formatting excel formula & inserting into shift list sheet
        #Check if next cell is empty in name column, inserts name & shift info in adjacent cell-column
        r=1
        while True:
                r +=1
                cell_name = sheet.cell(row=r,column=1)
                create_cell_border(cell_name)
                if cell_name.value is None: 
                        cell_name.value = "="+sheet_call+vol_cell.coordinate
                        cell_shift = sheet.cell(row = cell_name.row, column = cell_name.column+1)
                        cell_shift.value = "="+sheet_call+str(cells[0])+space_formula+sheet_call+str(cells[1])+space_formula+sheet_call+str(position_merge)+space_formula+sheet_call+str(time_value)
                        create_cell_border(cell_shift)
                        break

#Purpose: Helper function to create cell borders, given excel cell
def create_cell_border(cell):
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        cell.border = thin_border

#Purpose: Helper funtion that returns either 0 or 1, if a given cell is within the list of merged cells in a given sheet
def is_merged(cell,sheet):
        merged = 0
        for mergedCell in sheet.merged_cells.ranges:
                if cell.coordinate in mergedCell:
                        merged = 1
                        break
        return merged

#Purpose: Helper function that returns a Excel spreadsheet given a desired sheet title and a Excel workbook object
def create_shift_sheet(title,wb):
        if title in wb.sheetnames: sheet = wb[title]
        else:
                wb.create_sheet(title)
                sheet = wb[title]
        return sheet

#Main Executed Function
# #if __name__ == '__main__':
print("Creating Venue Schedule Grid")
args = parse_args()
if args.OutputFile is None:
        filename = args.FileName
        wb = pyxl.load_workbook(filename, read_only = False, keep_vba=True)
else:
        filename = args.OutputFile
        wb = Workbook(filename)
        wb.save(filename)
        wb = pyxl.load_workbook(filename, read_only = False, keep_vba=True)
wb.save(filename)
if args.VenueName in wb.sheetnames: 
        print('Venue Name already used in spreadsheet, exiting generation process')
        exit()
else:
        wb.create_sheet(args.VenueName)
        sheet = wb[args.VenueName]
        
volunteer_list = create_shift_sheet("VolunteerShifts",wb)        
supervisor_list = create_shift_sheet("SupervisorShifts",wb)  
wb.save(filename)
pos_vols = [args.BGVols,args.FOHVols,args.GTVols,args.HospVols,args.MerchVols,args.SiteVols,args.SecVols,args.FirstAidVols,args.SetupTeardownVols,args.OfficeVols,args.SurveyVols]
pos_names = ['Beer Garden','Front of House','Green Team','Hospitality','Merchandise','Site Crew','Security', 'First Aid','Setup/Teardown', 'Office','Survey']
#Pass all variables into excel spreadsheet generator program
#Using Column A as base cells, for chart formatting i.e. all formatting for merging and insertion will be done in column A
#Make Venue Name & Shift List Headers
sheet['A1'] = args.VenueName
sheet['A1'].font = title_font
sheet['A1'].alignment = align
volunteer_list['A1'] = 'Volunteers'
volunteer_list['B1'] = 'Shifts'
supervisor_list['A1'] = 'Supervisors'
supervisor_list['B1'] = 'Shifts'
create_cell_border(sheet['A1'])
sheet.column_dimensions['A'].width = len(args.VenueName) #Figure out good cell width value
sheet.row_dimensions[1].height = 30
sheet.merge_cells(start_row=1, start_column=1,end_row=1,end_column=int(args.NumberofDays))
sheet['A1'].fill = PatternFill("solid", fgColor="FFC3A8")
start = 4
for i in range(int(args.NumberofDays)):
        #Counter keeps track of cell traversals (done via the conditional loops below) in real time, for each column, resetting once a new column is selected
        counter = 0
        #pos_index is an internal counter for pos_names
        pos_index = 0
        #Create Date Cell
        cell = sheet.cell(row=2,column=1+i, value = 'Insert Date Here')
        create_cell_border(cell)
        cell_date = cell.coordinate
        #Create Show Title Cell
        cell = sheet.cell(row=3,column=1+i, value= 'Insert Band Name Here')
        cell.font = bolded_font
        create_cell_border(cell)
        #For every volunteer in a position, create a position, supervisor, and volunteer cells & insert 
        for vols in pos_vols:
                if int(vols) !=0:
                        #Position Name Cell
                        cell = sheet.cell(row=start+counter,column=1+i)
                        merged = is_merged(cell,sheet)
                        if merged == 0:
                                cell = sheet.cell(row=start+counter,column=1+i)
                                create_cell_border(cell)
                                cell.value = pos_names[pos_index] #Inserting Volunteer Position into cell
                                sheet.merge_cells(start_row=start+counter, start_column=1,end_row=start+counter,end_column=int(args.NumberofDays))
                                sheet.cell(row=start+counter,column=1+i).alignment = align
                                sheet.cell(row=start+counter,column=1+i).fill = PatternFill("solid", fgColor="FFC3A8")
                                cell_position = cell.coordinate
                        else: cell_position = cell.coordinate
                        split_shift = True
                        counter += 1
                        while True:
                                #Shift Time Cell
                                cell = sheet.cell(row=start+counter,column=1+i)
                                merged = is_merged(cell,sheet)
                                #Shift Time Cell Merge Requirement Check
                                if args.UniversalShiftTime == True and merged == 0:
                                        create_cell_border(cell)
                                        cell.value = args.ShiftTime
                                        sheet.merge_cells(start_row=start+counter, start_column=1,end_row=start+counter,end_column=int(args.NumberofDays))
                                if args.UniversalShiftTime == None:
                                        create_cell_border(cell)
                                        cell.value = args.ShiftTime
                                sheet.cell(row=start+counter,column=1+i).alignment = align
                                cell_time = cell.coordinate
                                #Supervisor Insertion
                                cell_venue = sheet.cell(row=1,column=1)
                                if args.SupervisorPositions is not None and pos_names[pos_index] in args.SupervisorPositions: 
                                        sup_cell = sheet.cell(row=start+counter+1,column=1+i,value = 'Insert supervisor name here')
                                        insert_sheet_list([cell_date,cell_venue.coordinate,cell_position,cell_time],supervisor_list,sheet,sup_cell)
                                        create_cell_border(sheet.cell(row=start+counter+1,column=1+i))
                                        sheet.cell(row=start+counter+1,column=1+i).fill = PatternFill("solid", fgColor="00FFCC99")
                                        sheet.cell(row=start+counter+1,column=1+i).font = supervisor_font
                                        counter +=1
                                #Volunteer Insertion
                                for j in range(int(vols)): 
                                        vol_cell = sheet.cell(row=start+counter+1+j,column=1+i,value = 'Insert volunteer name here')
                                        insert_sheet_list([cell_date,cell_venue.coordinate,cell_position,cell_time],volunteer_list,sheet,vol_cell)
                                        create_cell_border(sheet.cell(row=start+counter+1+j,column=1+i))
                                        sheet.cell(row=start+counter+1+j,column=1+i).fill = PatternFill("solid", fgColor="0099CC00")
                                        wb.save(filename)
                                counter +=int(vols)+1
                                #Conditional statements for split shift argument, breaks while loop if current volunteer position is split, or
                                #if no split shift is indicated for work position
                                if split_shift is False: break
                                if args.SplitShifts is None: break
                                if pos_names[pos_index] in args.SplitShifts: split_shift = False
                                else: break
                pos_index +=1
dims = {}
for row in sheet.rows:
        for cell in row:
                if cell.value: dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))  
        for col, value in dims.items(): sheet.column_dimensions[col].width = value
wb.save(filename)
print("Done")
